import numpy as np
import cv2
import time
import threading
import math
import datetime
import os
import openpyxl
import matplotlib
import openpyxl
import matplotlib
import pytesseract
from PIL import Image
import re
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import dlib
import matplotlib.pyplot as plt
from skimage import util
from skimage.io import imread
from skimage.filters import threshold_otsu
from skimage import measure
from skimage.measure import regionprops
import matplotlib.patches as patches
import sklearn
import joblib
count=0
papu=0
counter=0
SNo = int(1)
sheet = {}
now = datetime.datetime.now()
day = now.day
b = str(day) + '.' + str(now.month) + '.' + str(now.year)
c = datetime.datetime.now()
d = c.strftime('%d') + '/' + c.strftime('%b') + '/' + c.strftime('%Y')
e = c.strftime('%H') + ':' + c.strftime('%M') + ':' + c.strftime('%S')
mainFilename = b+ '.xlsx'

plate=cv2.CascadeClassifier('licenceplate.xml')
#plate = cv2.CascadeClassifier('cascade.xml')
#cap = cv2.VideoCapture('http://191.168.10.229/doc/page/login.asp?_1577432127346')#
#cap = cv2.VideoCapture('rtsp://admin:12345@192.168.60.24:554')
cap = cv2.VideoCapture('v1.mp4')

if not(os.path.isfile(mainFilename)):
    mwb = openpyxl.Workbook()
    main_sheet=mwb.active
    main_sheet.cell(row=1, column=1).value = 'VehicleID'
    main_sheet.cell(row=1, column=2).value = 'Vehicle_Number'
    sheet[0] = mwb.create_sheet('sheet%d' %counter)
    sheet[0] = mwb.active
    a = 'Daily Vehicles'
    day = now.day
    b = str(str(a) + str(day) + '.' + str(now.month) + '.' + str(now.year))
    sheet[0].title = b
    sheet[0].cell(row=1, column=1).value = 'Date'
    sheet[0].cell(row=1, column=2).value = 'Time'
    sheet[0].cell(row=1, column=3).value = 'Detected no'
    sheet[0].cell(row=1, column=4).value = 'Number Plate'
    mwb.save(mainFilename)
##################################        MODULES
def get_string(plate):
    # Read image with opencv
    # img = cv2.imread(img_path)
    img = cv2.cvtColor(plate, cv2.COLOR_BGR2GRAY)
    kernel = np.ones((1, 1), np.uint8)
    img = cv2.dilate(img, kernel, iterations=1)
    img = cv2.erode(img, kernel, iterations=1)
    result = pytesseract.image_to_string(img)
    # print(result)
    return result


def overlap(img, plate):
    plate = cv2.resize(plate, (400, 100), interpolation=cv2.INTER_NEAREST)
    heigh, wei = plate.shape[:2]
    #print(heigh, wei)

    while True:
        for i in range(0, heigh):
            for j in range(0, wei):
                img[1 + i, 800 + j] = plate[i, j]
        # cv2.imshow('overlay', img)
        # cv2.imshow('plt', plate)
        return (img)


def roiplt(count,a,b,c,d):
    plt = cv2.imread('./output/img' + str(count) + '.jpg')
    #print('./output/img'+str(count)+'.jpg')
    roiplt=plt[b+10:d-10, a+10:c-10]
    plate = roiplt
    if 1>0:
        no = '1' + get_string(plate)
        print(get_string(plate))
        count += 1
        if datetime.datetime.now().second % 1 == 0:
            if len(no) >= 10:
                cno = int(''.join(filter(str.isdigit, no)))
                if len(str(cno)) > 5:
                    print(no)
                    cv2.rectangle(img, (x, y), (x + w, y + h), (255, 0, 0), 2)
                    cv2.putText(img, "Status: {}".format('plate detected'), (30, 60), cv2.FONT_HERSHEY_SIMPLEX, 2,
                                (0, 0, 255), 3)
                    overlap(img, plate)
                    cv2.imwrite('./plt/img' + str(count) + '.jpg', roiplt)
                    c = datetime.datetime.now()
                    day = now.day
                    b = str(day) + '.' + str(now.month) + '.' + str(now.year)
                    d = c.strftime('%d') + '/' + c.strftime('%b') + '/' + c.strftime('%Y')
                    e = c.strftime('%H') + ':' + c.strftime('%M') + ':' + c.strftime('%S')
                    filename = b + '.xlsx'
                    mwb = load_workbook(filename)
                    sheet[counter] = mwb.active
                    max_row_sheet = sheet[counter].max_row
                    max_row_excel = int(max_row_sheet) + 1
                    sheet[counter].cell(row=max_row_excel, column=1).value = d
                    sheet[counter].cell(row=max_row_excel, column=2).value = e
                    sheet[counter].cell(row=max_row_excel, column=3).value = no
                    plateimg = 'plt/img%d.jpg' % count
                    pic = openpyxl.drawing.image.Image(plateimg)
                    sheet[counter].add_image(pic, 'D%d' % max_row_excel)
                    sheet[counter].row_dimensions[max_row_excel].height = 50
                    sheet[counter].column_dimensions['A'].width = 20
                    sheet[counter].column_dimensions['B'].width = 20
                    sheet[counter].column_dimensions['C'].width = 20
                    sheet[counter].column_dimensions['D'].width = 40
                    mwb.save(filename)
                else:
                    cv2.imwrite('./notplt/img' + str(count) + '.jpg', roiplt)
            else:
                cv2.imwrite('./notplt/img' + str(count) + '.jpg', roiplt)
        else:
            cv2.imwrite('./notplt/img' + str(count) + '.jpg', roiplt)

q=[0,0,0,0]
while 1:
    now = datetime.datetime.now()
    ret, img = cap.read()
    # img2=cv2.resize(img1, (1900, 1080), interpolation=cv2.INTER_NEAREST)
    # img = img2[50:1080,10:1900]
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    height,width=img.shape[:2]
    pt = plate.detectMultiScale(gray, 1.3, 5)
    for (x, y, w, h) in pt:
       # cv2.rectangle(img, (x, y), (x + w, y + h), (255,0,0),2)
        #cv2.putText(img, "Status: {}".format('plate detected'), (10, 20), cv2.FONT_HERSHEY_SIMPLEX,1, (0, 0, 255), 3)
        cv2.imwrite("output/img%d.jpg" % count, img)
        count += 1
        q[0]=x
        q[1]=y
        q[2]=int(x+w)
        q[3]=int(y+h)
        if datetime.datetime.now().second%2==0:
            roiplt(count-1, q[0], q[1], q[2], q[3])
        if count==1000:
            count=0
            continue

    cv2.imshow('img', img)
    k = cv2.waitKey(1) & 0xff
    if k == 27:
        break

cap.release()
cv2.destroyAllWindows()